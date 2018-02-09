import json
from openpyxl import load_workbook
import numpy as np


seps_dict={
    "seps": {
        "kpc": {
            "label": "KPC",
            "masked": 0,
            "maskflag": 0,
            "pres": 80.0,
            "results": {
                "qgas": 0.0,
                "qoil": 0.0,
                "qwat": 0.0
            }
        },
        "u2": {
            "label": "Unit-2",
            "masked": 0,
            "maskflag": 0,
            "pres": 80.0,
            "results": {
                "qgas": 0.0,
                "qoil": 0.0,
                "qwat": 0.0
            }
        },
        "u3": {
            "label": "Unit-3",
            "masked": 0,
            "maskflag": 0,
            "pres": 130.0,
            "results": {
                "qgas": 0.0,
                "qoil": 0.0,
                "qwat": 0.0
            }
        }
    }
}



conn_wb = load_workbook("well_connections.xlsm")
conns=conn_wb.get_sheet_by_name(name='wells')

gap={}
gap["wells"]={}
r=1
while conns.cell(row=r, column=1).value!=None:

    well=str(conns.cell(row=r, column=1).value)
    print(well)
    well_dict={

            "constraints": {
                "fwhp_min": -1,
                "qgas_max":-1
                },
            "gor": np.random.uniform(500.0,1500.0),
            "label": well,
            "masked": 0,
            "maskflag": 0,
            "dp_calc": 1,
            "pc": {
                "fwhps": [
                    1.01,
                    250.0
                ],
                "qoil": [
                    np.random.uniform(500.0,1500.0),
                    0.0
                ]
            },
            "results": {
                "dp": 0.0,
                "fwhp": 0.0,
                "pres": 0.0,
                "qgas": 0.0,
                "qoil": 0.0,
                "qwat": 0.0
            },
            "wct": np.random.uniform(0,20)*1.0

    }

    gap["wells"][well]=well_dict
    gap["wells"][well]["label"]=well
    r+=1


gap["joints"]={}
r=1
while conns.cell(row=r, column=2).value!=None:

    joint=str(conns.cell(row=r, column=2).value)

    joint_dict={
        "label": joint,
        "maskflag": 0,
        "results": {
            "pres": 0.0,
            "qgas": 0.0,
            "qoil": 0.0,
            "qwat": 0.0
        }
    }

    gap["joints"][joint]=joint_dict

    r+=1


gap["seps"]=seps_dict["seps"]


conns=conn_wb.get_sheet_by_name(name='pipes')
gap["pipes"]={}
r=1
while conns.cell(row=r, column=1).value!=None:
    to_item=str(conns.cell(row=r, column=1).value)
    from_item=str(conns.cell(row=r, column=2).value)
    pipe=str(conns.cell(row=r, column=3).value)
    print("===1",to_item,from_item,pipe)

    pipe_dict={
        "corr": {
            "a": 0.09,
            "b": 0.0,
            "type": "linear",
        },
        "from": from_item,
        "id": 4.5,
        "label": pipe,
        "lenght": 100.0,
        "masked": 0,
        "maskflag": 0,
        "results": {
            "dp": 0.0,
            "pres": 0.0,
            "qgas": 0.0,
            "qoil": 0.0,
            "qwat": 0.0
        },
        "to": to_item
    }

    gap["pipes"][pipe]=pipe_dict

    r+=1

r=1
while conns.cell(row=r, column=5).value!=None:
    to_item=str(conns.cell(row=r, column=5).value)
    from_item=str(conns.cell(row=r, column=6).value)
    pipe=str(conns.cell(row=r, column=7).value)
    print("===2",to_item,from_item,pipe)
    pipe_dict={
        "corr": {
            "a": 0.09,
            "b": 0.0,
            "type": "linear"
        },
        "from": from_item,
        "id": 4.5,
        "label": pipe,
        "lenght": 100.0,
        "masked": 0,
        "maskflag": 0,
        "results": {
            "dp": 0.0,
            "pres": 0.0,
            "qgas": 0.0,
            "qoil": 0.0,
            "qwat": 0.0
        },
        "to": to_item
    }

    gap["pipes"][pipe]=pipe_dict
    r+=1

r=1
while conns.cell(row=r, column=9).value!=None:
    to_item=str(conns.cell(row=r, column=9).value)
    from_item=str(conns.cell(row=r, column=10).value)
    pipe=str(conns.cell(row=r, column=11).value)
    masked=int(conns.cell(row=r, column=12).value)
    print("===3",to_item,from_item,pipe)
    pipe_dict={
        "corr": {
            "a": 0.09,
            "b": 0.0,
            "type": "linear"
        },
        "from": from_item,
        "id": 4.5,
        "label": pipe,
        "lenght": 100.0,
        "masked": masked,
        "maskflag": 0,
        "results": {
            "dp": 0.0,
            "pres": 0.0,
            "qgas": 0.0,
            "qoil": 0.0,
            "qwat": 0.0
        },
        "to": to_item
    }

    gap["pipes"][pipe]=pipe_dict
    r+=1



json.dump(gap, open("gap.json", 'w'),indent=4, sort_keys=True)
